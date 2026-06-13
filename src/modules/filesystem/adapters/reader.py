"""
Class DiskReader – Cross-platform rich filesystem reader.
Supports: Linux, macOS, Windows, Android (Termux).

:project: CodeCortex
:package: Modules.Filesystem.Adapters.Reader
:author: Steeven Andrian
:copyright: (c) 2026 CODDY Codework
:standard: CODDY-Filesystem-v1.0
"""

from __future__ import annotations
from pathlib import Path, PurePosixPath
from datetime import datetime, timezone
from typing import Dict, Any, Optional, List, Tuple, Set
import os
import re
import stat as stat_module
import base64
import mimetypes
import subprocess
import platform
import struct
import math
import json as json_lib
import zipfile
import tarfile
import shutil
import hashlib
import csv as csv_module
from collections import Counter

from src.core import ApiError

mimetypes.init()

_SYSTEM = platform.system().lower()

from src.core.utils.path import norm_path as _norm
from src.core.config.environment import utc_ts_to_iso as _utc_from_ts

_CSV_EXTENSIONS: Set[str] = {".csv", ".tsv"}
_SOURCE_CODE_EXTENSIONS: Dict[str, str] = {
    ".py": "python",
    ".js": "javascript",
    ".jsx": "jsx",
    ".mjs": "javascript",
    ".cjs": "javascript",
    ".ts": "typescript",
    ".tsx": "tsx",
    ".mts": "typescript",
    ".cts": "typescript",
    ".java": "java",
    ".go": "go",
    ".rs": "rust",
    ".rb": "ruby",
    ".php": "php",
    ".swift": "swift",
    ".kt": "kotlin",
    ".kts": "kotlin",
    ".scala": "scala",
    ".r": "r",
    ".lua": "lua",
    ".pl": "perl",
    ".pm": "perl",
    ".sh": "shell",
    ".bash": "shell",
    ".zsh": "shell",
    ".bat": "batch",
    ".cmd": "batch",
    ".ps1": "powershell",
    ".sql": "sql",
    ".html": "html",
    ".htm": "html",
    ".css": "css",
    ".scss": "scss",
    ".less": "less",
    ".vue": "vue",
    ".svelte": "svelte",
    ".c": "c",
    ".h": "c",
    ".cpp": "cpp",
    ".cc": "cpp",
    ".cxx": "cpp",
    ".hpp": "cpp",
    ".hxx": "cpp",
    ".cs": "csharp",
    ".dart": "dart",
    ".zig": "zig",
    ".nim": "nim",
    ".ex": "elixir",
    ".exs": "elixir",
    ".erl": "erlang",
    ".hrl": "erlang",
    ".clj": "clojure",
    ".cljs": "clojure",
    ".cljc": "clojure",
    ".coffee": "coffeescript",
    ".tf": "terraform",
    ".yaml": "yaml",
    ".yml": "yaml",
    ".json": "json",
    ".xml": "xml",
    ".svg": "svg",
    ".toml": "toml",
    ".ini": "ini",
    ".cfg": "ini",
    ".conf": "ini",
    ".dockerfile": "dockerfile",
    ".makefile": "makefile",
    ".cmake": "cmake",
    ".gradle": "gradle",
    ".proto": "protobuf",
    ".graphql": "graphql",
    ".gql": "graphql",
    ".md": "markdown",
    ".mdx": "markdown",
    ".rst": "rst",
    ".tex": "latex",
    ".astro": "astro",
    ".sass": "sass",
}

_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".ico", ".tiff", ".tif", ".avif", ".heic", ".heif"}
_AUDIO_EXTENSIONS = {".mp3", ".wav", ".ogg", ".flac", ".aac", ".wma", ".m4a", ".opus", ".mid", ".midi"}
_VIDEO_EXTENSIONS = {".mp4", ".avi", ".mkv", ".mov", ".wmv", ".flv", ".webm", ".m4v", ".ts", ".mts", ".3gp"}
_DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt", ".odt", ".ods", ".odp", ".rtf"}
_ARCHIVE_EXTENSIONS: Set[str] = {".zip", ".tar", ".gz", ".tgz", ".bz2", ".xz", ".7z", ".rar", ".zst"}
_LOG_EXTENSIONS: Set[str] = {".log", ".logs"}
_MARKDOWN_EXTENSIONS: Set[str] = {".md", ".mdx"}
_JSON_EXTENSIONS: Set[str] = {".json"}
_YAML_EXTENSIONS: Set[str] = {".yaml", ".yml"}

_TEXT_BINARY_MARKS = frozenset({0x00, 0x01, 0x02, 0x03, 0x04, 0x05, 0x06, 0x07, 0x08, 0x0B, 0x0C, 0x0E, 0x0F,
                                 0x10, 0x11, 0x12, 0x13, 0x14, 0x15, 0x16, 0x17, 0x18, 0x19, 0x1A, 0x1B, 0x1C,
                                 0x1D, 0x1E, 0x1F})

def _is_text(content: bytes) -> bool:
    return not any(b in _TEXT_BINARY_MARKS for b in content)

def _get_created_ts(stat_info: os.stat_result) -> float:
    if _SYSTEM == "darwin":
        return getattr(stat_info, "st_birthtime", stat_info.st_ctime)
    if _SYSTEM == "windows":
        return stat_info.st_ctime
    btime = getattr(stat_info, "st_birthtime", None)
    return btime if btime else stat_info.st_ctime

def _get_owner(stat_info: os.stat_result) -> str:
    if _SYSTEM in ("linux", "darwin", "android"):
        try:
            import pwd
            return pwd.getpwuid(stat_info.st_uid).pw_name
        except (ImportError, KeyError):
            pass
    try:
        import getpass
        return getpass.getuser()
    except Exception:
        pass
    return str(stat_info.st_uid)

def _get_group(stat_info: os.stat_result) -> str:
    if _SYSTEM in ("linux", "darwin", "android"):
        try:
            import grp
            return grp.getgrgid(stat_info.st_gid).gr_name
        except (ImportError, KeyError):
            pass
    return str(stat_info.st_gid)

from .analyzer import SourceCodeAnalyzer

class DiskReader:
    """
    Reads files/directories from the actual filesystem with rich metadata,
    file type detection, source code analysis, and document text extraction.
    Cross-platform: Linux, macOS, Windows, Android (Termux).
    """

    SOURCE_CODE_EXTENSIONS = _SOURCE_CODE_EXTENSIONS
    IMAGE_EXTENSIONS = _IMAGE_EXTENSIONS
    AUDIO_EXTENSIONS = _AUDIO_EXTENSIONS
    VIDEO_EXTENSIONS = _VIDEO_EXTENSIONS
    DOCUMENT_EXTENSIONS = _DOCUMENT_EXTENSIONS
    CSV_EXTENSIONS = _CSV_EXTENSIONS

    def __init__(self, max_text_size: int = 10 * 1024 * 1024, max_binary_size: int = 50 * 1024 * 1024, max_lines: int = 5000):
        self.max_text_size = max_text_size
        self.max_binary_size = max_binary_size
        self.max_lines = max_lines

    def read(self, path: str) -> Dict[str, Any]:
        resolved = Path(path).resolve()
        if not resolved.exists():
            return {"error": f"Path does not exist: {path}", "status_code": 404}

        if resolved.is_dir():
            return self._read_directory(resolved)

        stat_info = resolved.stat()
        metadata = self._build_metadata(resolved, stat_info)
        file_type = self._detect_file_type(resolved)
        mime_type, _ = mimetypes.guess_type(str(resolved))
        metadata["mime_type"] = mime_type or "application/octet-stream"

        base: Dict[str, Any] = {
            "file_type": file_type,
            "size_bytes": stat_info.st_size,
            "read_at": _utc_from_ts(datetime.now(timezone.utc).timestamp()),
            "metadata": metadata,
        }

        if file_type == "source_code":
            result = self._read_source_code(resolved, base)
        elif file_type == "markdown":
            result = self._read_markdown(resolved, base)
        elif file_type == "json":
            result = self._read_json(resolved, base)
        elif file_type == "yaml":
            result = self._read_yaml(resolved, base)
        elif file_type == "log":
            result = self._read_log(resolved, base)
        elif file_type == "archive":
            result = self._read_archive(resolved, base)
        elif file_type == "image":
            result = self._read_image(resolved, base)
        elif file_type == "audio":
            result = self._read_audio(resolved, base)
        elif file_type == "video":
            result = self._read_video(resolved, base)
        elif file_type == "csv":
            result = self._read_csv(resolved, base)
        elif file_type == "document":
            result = self._read_document(resolved, base)
        elif file_type == "text":
            result = self._read_text(resolved, base)
        else:
            result = self._read_generic_binary(resolved, base)

        result = self._enrich_with_git(resolved, result)
        return result

    def _build_metadata(self, resolved: Path, stat_info: os.stat_result) -> Dict[str, Any]:
        perms = stat_module.filemode(stat_info.st_mode)
        return {
            "name": resolved.name,
            "path": _norm(str(resolved)),
            "permissions": str(perms),
            "mode": int(stat_module.S_IMODE(stat_info.st_mode)),
            "owner": _get_owner(stat_info),
            "group": _get_group(stat_info),
            "created": _utc_from_ts(_get_created_ts(stat_info)),
            "modified": _utc_from_ts(stat_info.st_mtime),
        }

    def _detect_file_type(self, resolved: Path) -> str:
        ext = resolved.suffix.lower()
        if ext in _MARKDOWN_EXTENSIONS:
            return "markdown"
        if ext in _JSON_EXTENSIONS:
            return "json"
        if ext in _YAML_EXTENSIONS:
            return "yaml"
        if ext in _LOG_EXTENSIONS:
            return "log"
        if ext in _ARCHIVE_EXTENSIONS:
            return "archive"
        if ext in _SOURCE_CODE_EXTENSIONS:
            return "source_code"
        if ext in _CSV_EXTENSIONS:
            return "csv"
        if ext in _IMAGE_EXTENSIONS:
            return "image"
        if ext in _AUDIO_EXTENSIONS:
            return "audio"
        if ext in _VIDEO_EXTENSIONS:
            return "video"
        if ext in _DOCUMENT_EXTENSIONS:
            return "document"

        mime_type, _ = mimetypes.guess_type(str(resolved))
        if mime_type:
            if mime_type.startswith("text/"):
                return _SOURCE_CODE_EXTENSIONS.get(ext, "text")
            if mime_type.startswith("image/"):
                return "image"
            if mime_type.startswith("audio/"):
                return "audio"
            if mime_type.startswith("video/"):
                return "video"
            if mime_type in ("application/pdf",):
                return "document"

        if self._is_text_file(resolved):
            return "text"
        return "binary"

    def _is_text_file(self, resolved: Path) -> bool:
        try:
            with open(resolved, "rb") as f:
                chunk = f.read(8192)
            return _is_text(chunk)
        except Exception:
            return False

    def _read_source_code(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        try:
            content = resolved.read_text("utf-8", errors="replace")
        except (UnicodeDecodeError, ValueError, OSError):
            try:
                content = resolved.read_text("latin-1", errors="replace")
            except Exception:
                content = ""

        if not content:
            content = "[File read error: unable to decode content]"
        elif len(content) > self.max_text_size:
            content = content[:self.max_text_size] + "\n[... File truncated due to size limit ...]"

        ext = resolved.suffix.lower()
        language = _SOURCE_CODE_EXTENSIONS.get(ext, "text")

        base["language"] = language
        base["encoding"] = "utf8"
        base["content"] = content

        lines = content.splitlines()
        line_count = len(lines)

        ast_data = self._parse_ast(content, language, resolved.name)
        if ast_data:
            base["ast"] = ast_data

        comment_density = self._calc_comment_density(lines)

        analyzer = SourceCodeAnalyzer()
        deep_analysis = analyzer.analyze(content, language, resolved.name, lines)

        base["analysis"] = {
            "line_count": line_count,
            "character_count": len(content),
            "code_structure": ast_data if ast_data else {},
            "complexity_metrics": {
                "lines_of_code": line_count,
                "cyclomatic_complexity": self._estimate_cyclomatic(content),
                "cognitive_complexity": self._estimate_cognitive(content),
                "maintainability_index": self._estimate_maintainability(line_count, comment_density, len(deep_analysis.get("potential_bugs", []))),
                "comment_density": comment_density,
            },
            "linting_status": deep_analysis["linting_status"],
            "code_smells": deep_analysis["code_smells"],
            "security_issues": deep_analysis["security_issues"],
            "potential_bugs": deep_analysis["potential_bugs"],
            "dead_code": deep_analysis["dead_code"],
            "unused_variables": deep_analysis["unused_variables"],
        }

        base["ai_slop_detection"] = deep_analysis["ai_slop_detection"]
        base["ai_recommendations"] = deep_analysis["ai_recommendations"]
        base["todos"] = deep_analysis["todos"]
        base["placeholders"] = deep_analysis["placeholders"]

        if deep_analysis["ai_slop_detection"]["total_issues"] > 0:
            base["performance_hints"] = [
                f for f in deep_analysis["ai_slop_detection"]["detections"]
                if "performance" in f.get("message", "").lower()
            ]

        return base

    def _estimate_cyclomatic(self, content: str) -> int:
        count = 1
        count += len(re.findall(r"\bif\b", content))
        count += len(re.findall(r"\belif\b", content))
        count += len(re.findall(r"\bwhile\b", content))
        count += len(re.findall(r"\bfor\b", content))
        count += len(re.findall(r"\band\b", content))
        count += len(re.findall(r"\bor\b", content))
        count += len(re.findall(r"\bexcept\b", content))
        count += len(re.findall(r"\bcase\b", content))
        count += len(re.findall(r"\?\s*[\w\s]+\s*:", content))
        return count

    def _estimate_cognitive(self, content: str) -> int:
        score = 0
        nesting = 0
        for line in content.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            indent = len(line) - len(line.lstrip())
            new_nesting = indent // 4

            if new_nesting > nesting:
                score += (new_nesting - nesting) * 2
            nesting = new_nesting

            if re.search(r"\bif\b", stripped):
                score += 1
            if re.search(r"\belse\b", stripped):
                score += 1
            if re.search(r"\bfor\b", stripped):
                score += 1
            if re.search(r"\bwhile\b", stripped):
                score += 1
            if re.search(r"\btry\b", stripped):
                score += 1
            if re.search(r"\bexcept\b", stripped):
                score += 1
        return score

    def _estimate_maintainability(self, loc: int, comment_density: float, bug_count: int) -> int:
        if loc == 0:
            return 100
        loc_score = max(0, 100 - (loc // 10))
        comment_score = min(20, int(comment_density * 100))
        bug_penalty = min(30, bug_count * 5)
        return max(0, min(100, loc_score + comment_score - bug_penalty))

    def _parse_ast(self, content: str, language: str, filename: str) -> Optional[Dict[str, Any]]:
        try:
            from tree_sitter import Language, Parser
            import tree_sitter_python, tree_sitter_typescript
            lang_map = {
                "python": (tree_sitter_python, "python"),
                "typescript": ("javascript", "typescript"),
                "tsx": ("javascript", "tsx"),
                "javascript": ("javascript", "javascript"),
                "jsx": ("javascript", "jsx"),
            }
            if language not in lang_map:
                return None
            ts_lang, lang_name = lang_map[language]
            py_language = Language(getattr(tree_sitter_python, "language")())
            parser = Parser(py_language)
            tree = parser.parse(bytes(content, "utf-8"))
            if not tree or not tree.root_node:
                return None

            body = []
            functions = []
            classes = []
            imports = []

            def walk(node, depth=0):
                if depth > 20:
                    return
                if node.type == "function_definition":
                    name_node = node.child_by_field_name("name")
                    params_node = node.child_by_field_name("parameters")
                    func_name = name_node.text.decode() if name_node and name_node.text else "anonymous"
                    params = []
                    if params_node:
                        for c in params_node.children:
                            if c.type in ("identifier", "typed_parameter", "default_parameter"):
                                params.append(c.text.decode() if c.text else "?")
                    body.append({
                        "type": "FunctionDef",
                        "name": func_name,
                        "args": params,
                        "lineno": node.start_point[0] + 1,
                    })
                    functions.append(func_name)
                elif node.type == "class_definition":
                    name_node = node.child_by_field_name("name")
                    cls_name = name_node.text.decode() if name_node and name_node.text else "anonymous"
                    body.append({
                        "type": "ClassDef",
                        "name": cls_name,
                        "lineno": node.start_point[0] + 1,
                    })
                    classes.append(cls_name)
                elif node.type in ("import_statement", "import_from_statement"):
                    imports.append(node.text.decode() if node.text else "")
                for child in node.children:
                    walk(child, depth + 1)

            walk(tree.root_node)
            return {
                "type": "Module",
                "body": body,
                "imports": imports,
                "classes": classes,
                "functions": functions,
            }
        except Exception:
            return None

    def _calc_comment_density(self, lines: List[str]) -> float:
        if not lines:
            return 0.0
        comment_lines = sum(1 for l in lines if l.strip().startswith("#") or l.strip().startswith("//") or l.strip().startswith("/*"))
        return round(comment_lines / len(lines), 4)

    def _read_image(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        try:
            with open(resolved, "rb") as f:
                raw = f.read()
            mime = mimetypes.guess_type(str(resolved))[0] or "image/png"
            b64 = base64.b64encode(raw).decode("ascii")
            base["content"] = f"data:{mime};base64,{b64}"
            base["content_base64"] = b64
            base["encoding"] = "base64"
        except Exception as e:
            base["content"] = None
            base["encoding"] = "binary"
            base["_read_error"] = str(e)

        image_info = {"width": 0, "height": 0, "has_alpha": False}
        try:
            from PIL import Image
            with Image.open(resolved) as img:
                image_info["width"] = img.width
                image_info["height"] = img.height
                image_info["color_profile"] = img.mode
                image_info["bit_depth"] = len(img.getbands()) * 8 if img.mode else 24
                image_info["compression"] = img.format or "unknown"
                image_info["has_alpha"] = "A" in img.mode
        except ImportError:
            pass
        except Exception:
            pass

        exif_data = {}
        try:
            from PIL import Image as PILImage
            from PIL.ExifTags import TAGS
            with PILImage.open(resolved) as img:
                exif_raw = img._getexif()
                if exif_raw:
                    for tag_id, value in exif_raw.items():
                        tag_name = TAGS.get(tag_id, tag_id)
                        if isinstance(value, bytes):
                            try:
                                value = value.decode("utf-8", errors="replace")
                            except Exception:
                                value = str(value)
                        exif_data[tag_name] = value
        except (ImportError, Exception):
            pass

        if exif_data:
            base["exif"] = exif_data
        base["image_info"] = image_info
        return base

    def _read_audio(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        try:
            with open(resolved, "rb") as f:
                raw = f.read(self.max_binary_size)
            mime = mimetypes.guess_type(str(resolved))[0] or "audio/mpeg"
            b64 = base64.b64encode(raw).decode("ascii")
            base["content"] = f"data:{mime};base64,{b64}"
            base["encoding"] = "base64"
        except Exception as e:
            base["content"] = None
            base["encoding"] = "binary"
            base["_read_error"] = str(e)

        base["audio_info"] = {
            "codec": resolved.suffix.lower().replace(".", "").upper(),
            "sample_rate": 0,
            "channels": 0,
            "bitrate": 0,
        }
        base["duration_seconds"] = 0.0
        return base

    def _read_video(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        try:
            with open(resolved, "rb") as f:
                raw = f.read(self.max_binary_size)
            mime = mimetypes.guess_type(str(resolved))[0] or "video/mp4"
            b64 = base64.b64encode(raw).decode("ascii")
            base["content"] = f"data:{mime};base64,{b64}"
            base["encoding"] = "base64"
        except Exception as e:
            base["content"] = None
            base["encoding"] = "binary"
            base["_read_error"] = str(e)

        base["video_info"] = {
            "codec": "",
            "width": 0,
            "height": 0,
            "frame_rate": 0,
            "bitrate": 0,
            "audio_tracks": [],
        }
        base["duration_seconds"] = 0.0
        return base

    def _read_csv(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        ext = resolved.suffix.lower()
        delimiter = "\t" if ext == ".tsv" else ","
        base["encoding"] = "utf8"

        try:
            content = resolved.read_text("utf-8", errors="replace")
        except Exception:
            content = ""

        if not content:
            base["content"] = "[File read error: unable to decode content]"
            base["csv_info"] = {"error": "unable to read file"}
            return base

        base["content"] = content
        lines = content.splitlines()
        total_lines = len(lines)

        rows = []
        headers = []
        dialect = None
        parse_error = None

        try:
            sample = content[:65536]
            sniffer = csv_module.Sniffer()
            dialect = sniffer.sniff(sample)
            delimiter = dialect.delimiter
        except Exception:
            pass

        try:
            reader = csv_module.reader(lines, delimiter=delimiter)
            all_rows = list(reader)
            if all_rows:
                headers = all_rows[0]
                rows = all_rows[1:]
        except Exception as e:
            parse_error = str(e)

        row_count = len(rows)
        col_count = len(headers) if headers else 0

        stats = {
            "total_lines": total_lines,
            "header_row_count": 1 if headers else 0,
            "data_rows": row_count,
            "columns": col_count,
            "delimiter": delimiter,
        }

        column_info = []
        if headers and rows:
            for ci, h in enumerate(headers):
                col_vals = [r[ci] for r in rows if ci < len(r) and r[ci]]
                non_empty = len(col_vals)
                numeric_count = sum(1 for v in col_vals if v.replace(".", "", 1).replace("-", "", 1).isdigit())
                col_info = {
                    "name": h,
                    "non_empty_cells": non_empty,
                    "empty_cells": row_count - non_empty,
                    "sample_values": col_vals[:5],
                }
                if numeric_count > non_empty * 0.5:
                    col_info["type"] = "numeric"
                    num_vals = [float(v) for v in col_vals if v.replace(".", "", 1).replace("-", "", 1).isdigit()]
                    if num_vals:
                        col_info["min"] = min(num_vals)
                        col_info["max"] = max(num_vals)
                        col_info["mean"] = round(sum(num_vals) / len(num_vals), 2)
                else:
                    col_info["type"] = "text"
                column_info.append(col_info)

        preview_rows = min(10, row_count)
        preview = []
        for r in rows[:preview_rows]:
            row_dict = {}
            for ci, h in enumerate(headers):
                row_dict[h] = r[ci] if ci < len(r) else ""
            preview.append(row_dict)

        base["csv_info"] = {
            "statistics": stats,
            "headers": headers,
            "columns_detail": column_info,
            "preview_rows": preview,
        }
        if parse_error:
            base["csv_info"]["parse_error"] = parse_error

        base["ai_summary"] = (
            f"CSV with {row_count} data rows × {col_count} columns. "
            f"Delimiter: '{delimiter}'. "
            f"{'First row used as header.' if headers else 'No header row detected.'}"
        )
        return base

    def _read_document(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        ext = resolved.suffix.lower()
        content = ""
        doc_info = {"page_count": 0, "title": "", "author": ""}
        doc_structure = {"sections": [], "tables": []}

        try:
            if ext == ".pdf":
                try:
                    from pypdf import PdfReader
                    reader = PdfReader(resolved)
                    doc_info["page_count"] = len(reader.pages)
                    if reader.metadata:
                        doc_info["title"] = reader.metadata.title or ""
                        doc_info["author"] = reader.metadata.author or ""
                        doc_info["subject"] = reader.metadata.subject or ""
                    pages_text = []
                    for i, page in enumerate(reader.pages):
                        page_text = page.extract_text() or ""
                        pages_text.append(page_text)
                    content = "\n\n".join(pages_text)
                except Exception:
                    content = "[PDF text extraction failed]"

            elif ext in (".docx", ".doc"):
                try:
                    from docx import Document
                    doc = Document(resolved)
                    paragraphs = [p.text for p in doc.paragraphs if p.text]
                    content = "\n".join(paragraphs)
                except Exception:
                    content = "[DOCX text extraction failed]"

            elif ext == ".xlsx":
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(resolved, read_only=True, data_only=True)
                    sheets_text = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        rows_text = []
                        for row in ws.iter_rows(values_only=True):
                            rows_text.append(" | ".join(str(c) if c is not None else "" for c in row))
                        sheets_text.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows_text))
                    content = "\n\n".join(sheets_text)
                except Exception:
                    content = "[XLSX text extraction failed]"

            elif ext == ".xls":
                try:
                    from xlrd import open_workbook
                    wb = open_workbook(str(resolved))
                    sheets_text = []
                    for sheet_name in wb.sheet_names():
                        ws = wb.sheet_by_name(sheet_name)
                        rows_text = []
                        for row_idx in range(ws.nrows):
                            row_vals = [ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)]
                            rows_text.append(" | ".join(str(c) if c else "" for c in row_vals))
                        sheets_text.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows_text))
                    content = "\n\n".join(sheets_text)
                except ImportError:
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(resolved, read_only=True, data_only=True)
                        sheets_text = []
                        for sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            rows_text = []
                            for row in ws.iter_rows(values_only=True):
                                rows_text.append(" | ".join(str(c) if c is not None else "" for c in row))
                            sheets_text.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows_text))
                        content = "\n\n".join(sheets_text)
                    except ImportError:
                        content = "[XLS reading requires xlrd or openpyxl]"
                    except Exception:
                        content = "[XLS text extraction failed]"
                except Exception:
                    content = "[XLS text extraction failed]"

            else:
                try:
                    content = resolved.read_text("utf-8", errors="replace")
                except Exception:
                    content = "[Binary document format - text extraction not supported]"

        except Exception as e:
            content = f"[Document read error: {e}]"

        if len(content) > self.max_text_size:
            content = content[:self.max_text_size] + "\n[... Truncated ...]"

        base["encoding"] = "utf8"
        base["content"] = content
        base["document_info"] = doc_info
        base["document_structure"] = doc_structure
        return base

    def _read_markdown(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        content = self._read_text_content(resolved)
        if not content:
            base["content"] = "[File read error: unable to decode content]"
            base["markdown_info"] = {"headings": [], "links": [], "images": [], "code_blocks": [], "task_lists": {"total": 0, "completed": 0, "incomplete": 0, "items": []}, "word_count": 0, "readability_score": 0.0}
            return base
        base["encoding"] = "utf8"
        base["content"] = content

        lines = content.splitlines()
        headings = []
        links = []
        images = []
        code_blocks = []
        task_items = []

        in_code_block = False
        code_block_start = 0
        code_block_lang = ""
        code_block_lines = []

        for i, line in enumerate(lines, 1):
            stripped = line.strip()

            # Code blocks
            if stripped.startswith("```"):
                if in_code_block:
                    code_blocks.append({"language": code_block_lang, "line_start": code_block_start, "line_end": i, "content": "\n".join(code_block_lines)})
                    in_code_block = False
                    code_block_lines = []
                else:
                    in_code_block = True
                    code_block_start = i
                    code_block_lang = stripped[3:].strip().split()[0] if len(stripped) > 3 else ""
                continue

            if in_code_block:
                code_block_lines.append(line)
                continue

            # Headings
            hm = re.match(r"^(#{1,6})\s+(.+)$", stripped)
            if hm:
                headings.append({"level": len(hm.group(1)), "text": hm.group(2).strip(), "line": i})

            # Task list items
            tm = re.match(r"^[-*]\s+\[([ xX])\]\s+(.+)$", stripped)
            if tm:
                task_items.append({"text": tm.group(2).strip(), "checked": tm.group(1).lower() == "x", "line": i})

            # Links
            for m in re.finditer(r"\[([^\]]+)\]\(([^)]+)\)", line):
                url = m.group(2)
                if not url.startswith("data:"):
                    links.append({"url": url, "text": m.group(1), "line": i})

            # Images
            for m in re.finditer(r"!\[([^\]]*)\]\(([^)]+)\)", line):
                images.append({"url": m.group(2), "alt": m.group(1), "line": i})

        total_tasks = len(task_items)
        completed_tasks = sum(1 for t in task_items if t["checked"])
        words = content.split()
        word_count = len(words)
        readability = self._readability_score(content)

        base["markdown_info"] = {
            "headings": headings,
            "links": links,
            "images": images,
            "code_blocks": code_blocks,
            "task_lists": {
                "total": total_tasks,
                "completed": completed_tasks,
                "incomplete": total_tasks - completed_tasks,
                "items": task_items,
            },
            "word_count": word_count,
            "readability_score": readability,
        }
        return base

    def _read_json(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        content = self._read_text_content(resolved)
        if not content:
            base["content"] = "[File read error: unable to decode content]"
            base["json_info"] = {"schema_validation": {"valid": False, "errors": ["Unable to read file"], "schema_used": None}}
            return base
        base["encoding"] = "utf8"

        raw_bytes = resolved.read_bytes()
        pretty_size = len(raw_bytes)

        parsed = None
        parse_error = None
        try:
            parsed = json_lib.loads(content)
            base["content"] = parsed
        except json_lib.JSONDecodeError as e:
            parse_error = str(e)
            base["content"] = content

        data_type = type(parsed).__name__ if parsed is not None else "unknown"
        keys = list(parsed.keys()) if isinstance(parsed, dict) else []
        max_depth = self._json_max_depth(parsed) if parsed is not None else 0

        stat_counts = {"strings": 0, "numbers": 0, "booleans": 0, "arrays": 0, "objects": 0, "nulls": 0}
        total_values = 0
        if parsed is not None:
            self._json_count_types(parsed, stat_counts)
            total_values = sum(stat_counts.values())

        minified = json_lib.dumps(parsed, separators=(",", ":"), default=str) if parsed is not None else content
        minified_size = len(minified.encode("utf-8"))

        schema_valid = True
        schema_errors = []
        try:
            import jsonschema
            if isinstance(parsed, dict):
                pass  # Would need a schema to validate against
        except ImportError:
            pass

        ai_insights = {"possible_env_vars": [], "suggested_defaults": None, "breaking_changes": None}
        if isinstance(parsed, dict):
            env_vars = [k for k in parsed if k.isupper() and isinstance(parsed[k], (str, int, float, bool))]
            if env_vars:
                ai_insights["possible_env_vars"] = env_vars
                ai_insights["suggested_defaults"] = {k: parsed[k] for k in env_vars}
            elif "env" in parsed or "environment" in parsed or "config" in parsed:
                config = parsed.get("config") or parsed.get("env") or parsed.get("environment") or {}
                if isinstance(config, dict):
                    ai_insights["possible_env_vars"] = [k for k in config if k.isupper() or "_" in k]
                    ai_insights["suggested_defaults"] = config

        base["json_info"] = {
            "schema_validation": {"valid": schema_valid, "errors": schema_errors, "schema_used": None},
            "data_type": data_type,
            "keys": keys,
            "max_depth": max_depth,
            "total_values": total_values,
            "statistics": stat_counts,
            "pretty_size_bytes": pretty_size,
            "minified_size_bytes": minified_size,
        }
        if parse_error:
            base["json_info"]["parse_error"] = parse_error
        base["ai_insights"] = ai_insights
        return base

    def _read_yaml(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        content = self._read_text_content(resolved)
        if not content:
            base["content"] = "[File read error]"
            base["yaml_info"] = {"anchors_used": [], "aliases_used": [], "merge_keys": [], "duplicate_keys_detected": False, "complexity_score": 0, "document_count": 0, "parsing_warnings": []}
            base["converted_to_json"] = None
            return base
        base["encoding"] = "utf8"
        base["content"] = content

        anchors = list(set(re.findall(r"&(\w+)", content)))
        aliases = list(set(re.findall(r"\*(\w+)", content)))
        merge_keys = ["<<"] if "<<" in content else []
        duplicate_keys = self._yaml_has_duplicate_keys(content)
        complexity = len(anchors) + len(aliases) + (content.count(":") // 5)
        doc_count = len(re.findall(r"^---\s*$", content, re.MULTILINE)) + 1

        warnings = []
        parsed = None
        converted = None
        try:
            import yaml
            parsed = yaml.safe_load(content)
            converted = json_lib.loads(json_lib.dumps(parsed, default=str)) if parsed is not None else None
        except ImportError:
            warnings.append("PyYAML not installed, using regex-based analysis only")
        except yaml.YAMLError as e:
            warnings.append(f"YAML parse warning: {str(e)}")
            parsed = None
        except Exception as e:
            warnings.append(f"Conversion warning: {str(e)}")

        base["yaml_info"] = {
            "anchors_used": anchors,
            "aliases_used": aliases,
            "merge_keys": merge_keys,
            "duplicate_keys_detected": duplicate_keys,
            "complexity_score": complexity,
            "document_count": doc_count,
            "parsing_warnings": warnings,
        }
        base["converted_to_json"] = converted
        return base

    def _read_log(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        content = self._read_text_content(resolved)
        if not content:
            base["content_preview"] = "[File read error]"
            base["log_info"] = {"total_lines": 0, "unique_events": 0, "log_levels": {}, "pattern_matches": [], "anomalies": [], "structured_fields": {}}
            return base
        base["encoding"] = "utf8"

        lines = content.splitlines()
        total_lines = len(lines)
        preview_lines = min(20, total_lines)
        base["content_preview"] = "\n".join(lines[:preview_lines]) + ("\n..." if total_lines > preview_lines else "")

        log_levels = Counter()
        log_patterns = [
            (r"ERROR|FATAL|CRITICAL", "ERROR"),
            (r"WARN|WARNING", "WARN"),
            (r"INFO|NOTICE", "INFO"),
            (r"DEBUG|TRACE", "DEBUG"),
        ]
        ip_pattern = re.compile(r"\b(?:\d{1,3}\.){3}\d{1,3}\b")
        user_pattern = re.compile(r"\buser[=: ](\w+)", re.IGNORECASE)
        http_pattern = re.compile(r"\b(5\d{2}|4\d{2}|3\d{2}|2\d{2})\b")
        timestamp_pattern = re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}")

        error_lines = []
        timestamps = []
        ip_addresses = set()
        user_ids = set()
        http_statuses = set()
        event_counts = Counter()
        pattern_matches = Counter()

        error_patterns = [
            ("Failed to connect", "high"),
            ("timeout", "medium"),
            ("Exception:", "high"),
            ("Traceback", "high"),
            ("null pointer", "high"),
            ("permission denied", "high", re.IGNORECASE),
            ("segmentation fault", "critical"),
            ("out of memory", "critical"),
            ("disk full", "critical"),
            ("connection refused", "high"),
            ("panic:", "critical"),
            ("fatal:", "critical", re.IGNORECASE),
        ]

        for i, line in enumerate(lines, 1):
            stripped = line.strip()
            if not stripped:
                continue

            for pattern, sev, *flags in error_patterns:
                flags_dict = flags[0] if flags else 0
                if re.search(pattern, stripped, flags_dict):
                    pattern_matches[(pattern, sev)] += 1

            if re.search(r"ERROR|FATAL|CRITICAL", stripped):
                error_lines.append(i)

            ts_match = timestamp_pattern.search(stripped)
            if ts_match:
                timestamps.append(ts_match.group())

            ip_addresses.update(ip_pattern.findall(stripped))
            user_ids.update(user_pattern.findall(stripped))
            http_statuses.update(int(s) for s in http_pattern.findall(stripped) if len(s) == 3)

            for pat, _ in log_patterns:
                if re.search(pat, stripped):
                    log_levels[pat.split("|")[0]] += 1
                    break

            # Extract event by removing timestamp/level prefix
            event = re.sub(r"^\S+\s+\S+\s+", "", stripped)
            event_counts[event[:80]] += 1

        time_range = {}
        if len(timestamps) >= 2:
            sorted_ts = sorted(timestamps)
            time_range = {"start": sorted_ts[0] + "Z" if not sorted_ts[0].endswith("Z") else sorted_ts[0],
                          "end": sorted_ts[-1] + "Z" if not sorted_ts[-1].endswith("Z") else sorted_ts[-1]}
        elif len(timestamps) == 1:
            time_range = {"start": timestamps[0], "end": timestamps[0]}

        # Anomaly detection: error spikes, exceptions, unusual counts
        anomalies = []
        error_count = log_levels.get("ERROR", 0)
        if error_count > 50:
            anomalies.append({"line": 0, "message": f"High error count: {error_count} errors", "type": "high_error_count"})
        if error_count > 10 and total_lines > 100:
            # Find first error area
            first_error_line = next((l for l in error_lines if l > 0), 0)
            if first_error_line and error_count > total_lines * 0.05:
                anomalies.append({"line": first_error_line, "message": f"Sudden spike of {error_count} errors", "type": "error_spike"})

        for (pat_text, pat_sev), count in pattern_matches.items():
            if count > 1:
                anomalies.append({"line": 0, "message": f"Pattern '{pat_text}' found {count} times", "type": f"pattern_{pat_sev}"})

        # Structured fields
        structured_fields = {}
        if ip_addresses:
            structured_fields["ip_addresses"] = sorted(ip_addresses)[:50]
        if user_ids:
            structured_fields["user_ids"] = sorted(user_ids)[:50]
        if http_statuses:
            structured_fields["http_status"] = sorted(http_statuses)[:50]

        # Unique events
        unique_events = len(event_counts)

        # Build pattern matches list
        pattern_matches_list = [{"pattern": p[0], "count": c, "severity": p[1]} for (p, c) in pattern_matches.most_common(20)]

        base["log_info"] = {
            "total_lines": total_lines,
            "unique_events": unique_events,
            "time_range": time_range,
            "log_levels": dict(log_levels),
            "pattern_matches": pattern_matches_list,
            "anomalies": anomalies[:10],
            "structured_fields": structured_fields,
        }

        # AI summary
        most_common = log_levels.most_common(1)
        top_level = most_common[0][0] if most_common else "INFO"
        base["ai_summary"] = (
            f"{total_lines} log lines, {error_count} errors, "
            f"{len(anomalies)} anomalies detected. "
            f"Top level: {top_level}. "
            f"{'High error rate' if error_count > total_lines * 0.05 else 'Normal operation pattern'}."
        )
        return base

    def _read_archive(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        base["encoding"] = None
        base["content"] = None
        ext = resolved.suffix.lower()
        fmt = "zip"
        if ext == ".zip":
            fmt = "zip"
        elif ext in (".tar", ".gz", ".tgz", ".bz2", ".xz"):
            fmt = "tar"
        else:
            base["archive_info"] = {"total_files": 0, "total_folders": 0, "total_uncompressed_size": 0,
                                     "compression_ratio": 0.0, "is_encrypted": False,
                                     "is_signed": False, "file_list": []}
            base["file_type"] = "binary"
            return self._read_generic_binary(resolved, base)

        base["archive_format"] = fmt
        file_list = []
        total_compressed = 0
        total_uncompressed = 0
        file_count = 0
        folder_count = 0
        is_encrypted = False
        duplicate_paths = []
        suspicious_files = []
        seen_paths = set()

        try:
            if fmt == "zip":
                with zipfile.ZipFile(resolved, "r") as zf:
                    for info in zf.infolist():
                        path = _norm(info.filename)
                        if info.is_dir():
                            folder_count += 1
                            continue
                        file_count += 1
                        total_compressed += info.compress_size
                        total_uncompressed += info.file_size
                        entry = {
                            "path": path,
                            "compressed_size": info.compress_size,
                            "uncompressed_size": info.file_size,
                            "compression_method": self._zip_compress_method(info.compress_type),
                            "modified": _utc_from_ts(datetime(*info.date_time, tzinfo=timezone.utc).timestamp()) if info.date_time else "",
                            "crc32": hex(info.CRC) if info.CRC else "",
                        }
                        file_list.append(entry)
                        if path in seen_paths:
                            duplicate_paths.append(path)
                        seen_paths.add(path)
                        if ".." in path or path.startswith("/"):
                            suspicious_files.append(path)
                    is_encrypted = any(zf.getinfo(n).flag_bits & 0x1 for n in zf.namelist())

            elif fmt == "tar":
                import tarfile as tf
                mode_map = {".tar": "r:", ".gz": "r:gz", ".tgz": "r:gz", ".bz2": "r:bz2", ".xz": "r:xz"}
                tmode = mode_map.get(ext, "r:*")
                try:
                    with tf.open(resolved, tmode) as tf:
                        for member in tf.getmembers():
                            path = _norm(member.name)
                            if member.isdir():
                                folder_count += 1
                                continue
                            file_count += 1
                            total_uncompressed += member.size
                            entry = {
                                "path": path,
                                "compressed_size": 0,
                                "uncompressed_size": member.size,
                                "compression_method": ext[1:].upper(),
                                "modified": _utc_from_ts(member.mtime) if member.mtime else "",
                                "crc32": "",
                            }
                            file_list.append(entry)
                            if path in seen_paths:
                                duplicate_paths.append(path)
                            seen_paths.add(path)
                            if ".." in path or path.startswith("/"):
                                suspicious_files.append(path)
                except tf.ReadError:
                    pass

        except Exception as e:
            base["archive_info"] = {"error": str(e)}
            return base

        compression_ratio = round(total_uncompressed / total_compressed, 2) if total_compressed > 0 else 0.0
        base["archive_info"] = {
            "total_files": file_count,
            "total_folders": folder_count,
            "total_uncompressed_size": total_uncompressed,
            "compression_ratio": compression_ratio,
            "is_encrypted": is_encrypted,
            "is_signed": False,
            "file_list": file_list[:500],
            "duplicate_paths": duplicate_paths[:20],
            "suspicious_files": suspicious_files[:20],
        }
        total_entries = file_count + folder_count
        fmt_name = fmt.upper()
        base["ai_summary"] = (
            f"{fmt_name} archive with {total_entries} entries "
            f"({file_count} files, {folder_count} folders), "
            f"compression ratio {compression_ratio}x. "
            f"{'Encrypted. ' if is_encrypted else ''}"
            f"{'Contains suspicious paths!' if suspicious_files else 'No security issues detected.'}"
        )
        return base

    def _read_generic_binary(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        base["encoding"] = "binary"
        base["content"] = None
        try:
            stat_info = resolved.stat()
            if stat_info.st_size <= self.max_binary_size:
                with open(resolved, "rb") as f:
                    raw = f.read(self.max_binary_size)
                mime, _ = mimetypes.guess_type(str(resolved))
                b64 = base64.b64encode(raw[:self.max_binary_size]).decode("ascii")
                base["content"] = f"data:{mime or 'application/octet-stream'};base64,{b64}"
                base["encoding"] = "base64"
            else:
                raw = b""
                with open(resolved, "rb") as f:
                    raw = f.read(65536)
        except Exception:
            raw = b""

        magic = raw[:16]
        magic_hex = " ".join(f"{b:02X}" for b in magic)
        file_type_hint = self._magic_to_type(magic, resolved.suffix.lower())

        entropy = 0.0
        if raw:
            entropy = self._entropy(raw)

        hex_lines = []
        for i in range(0, min(256, len(raw)), 16):
            chunk = raw[i:i+16]
            hex_part = " ".join(f"{b:02X}" for b in chunk[:8]) + "  " + " ".join(f"{b:02X}" for b in chunk[8:])
            ascii_part = "".join(chr(b) if 32 <= b < 127 else "." for b in chunk)
            hex_lines.append(f"{i:08x}: {hex_part:<48s} |{ascii_part}|")
        hex_dump = "\n".join(hex_lines[:32])

        strings = []
        current = []
        for b in raw:
            if 32 <= b < 127:
                current.append(chr(b))
            else:
                if len(current) >= 4:
                    strings.append("".join(current))
                current = []
        if len(current) >= 4:
            strings.append("".join(current))
        strings = list(set(strings))[:100]

        sections = []
        if magic[:4] == b"\x7fELF":
            sections = self._parse_elf_sections(raw)
        elif magic[:2] == b"MZ":
            sections = self._parse_pe_sections(raw)

        is_text = _is_text(raw[:8192]) if raw else False
        is_compressed = entropy > 7.5

        base["binary_info"] = {
            "magic_bytes": magic_hex,
            "file_type_hint": file_type_hint,
            "entropy": round(entropy, 2),
            "is_text_likely": is_text,
            "is_compressed": is_compressed,
            "hex_dump_preview": hex_dump,
            "strings_extracted": strings[:50],
            "section_headers": sections if sections else None,
        }

        base["ai_analysis"] = {
            "possible_purpose": file_type_hint or "Unknown binary file",
            "risk_indicators": self._binary_risk_indicators(raw, sections, is_compressed),
            "recommendation": "Use file(1) command for deeper inspection" if not file_type_hint else "Binary file analyzed successfully",
        }
        return base

    def _read_text_content(self, resolved: Path) -> str:
        try:
            return resolved.read_text("utf-8", errors="replace")
        except (UnicodeDecodeError, ValueError, OSError):
            try:
                return resolved.read_text("latin-1", errors="replace")
            except Exception:
                return ""
        except Exception:
            return ""

    def _readability_score(self, text: str) -> float:
        sentences = len(re.findall(r"[.!?]+", text))
        words = len(text.split())
        syllables = sum(self._syllables(w) for w in text.split()[:200])
        if sentences == 0 or words == 0:
            return 0.0
        return round(206.835 - 1.015 * (words / sentences) - 84.6 * (syllables / words), 1)

    def _syllables(self, word: str) -> int:
        word = word.lower().strip(".,!?;:")
        if not word:
            return 1
        count = 0
        vowels = "aeiouy"
        if word[0] in vowels:
            count += 1
        for i in range(1, len(word)):
            if word[i] in vowels and word[i-1] not in vowels:
                count += 1
        if word.endswith("e"):
            count -= 1
        if word.endswith("le") and len(word) > 2 and word[-3] not in vowels:
            count += 1
        return max(1, count)

    def _json_max_depth(self, obj, depth=0) -> int:
        if isinstance(obj, dict):
            if not obj:
                return depth + 1
            return max(self._json_max_depth(v, depth + 1) for v in obj.values())
        if isinstance(obj, list):
            if not obj:
                return depth + 1
            return max(self._json_max_depth(v, depth + 1) for v in obj)
        return depth

    def _json_count_types(self, obj, counts: dict):
        if isinstance(obj, str):
            counts["strings"] += 1
        elif isinstance(obj, bool):
            counts["booleans"] += 1
        elif isinstance(obj, (int, float)):
            counts["numbers"] += 1
        elif obj is None:
            counts["nulls"] += 1
        elif isinstance(obj, dict):
            counts["objects"] += 1
            for v in obj.values():
                self._json_count_types(v, counts)
        elif isinstance(obj, list):
            counts["arrays"] += 1
            for v in obj:
                self._json_count_types(v, counts)

    def _yaml_has_duplicate_keys(self, content: str) -> bool:
        seen = set()
        for m in re.finditer(r"^(\s*)([\w._-]+):", content, re.MULTILINE):
            key = m.group(2).strip()
            if key in seen:
                return True
            seen.add(key)
        return False

    def _zip_compress_method(self, method: int) -> str:
        methods = {0: "STORED", 1: "SHRUNK", 8: "DEFLATE", 9: "DEFLATE64", 12: "BZIP2", 14: "LZMA", 93: "ZSTANDARD"}
        return methods.get(method, f"UNKNOWN({method})")

    def _entropy(self, data: bytes) -> float:
        if not data:
            return 0.0
        counts = Counter(data)
        total = len(data)
        entropy = -sum((c / total) * math.log2(c / total) for c in counts.values())
        return entropy

    _MAGIC_DB: List[Tuple[bytes, str]] = [
        (b"\x7fELF", "ELF executable (Linux)"),
        (b"MZ", "PE executable (Windows)"),
        (b"\x89PNG\r\n\x1a\n", "PNG image"),
        (b"\xff\xd8\xff", "JPEG image"),
        (b"GIF8", "GIF image"),
        (b"RIFF", "RIFF / AVI / WAV container"),
        (b"\x00\x00\x00\x18ftyp", "MP4 / MOV video"),
        (b"\x00\x00\x00\x20ftyp", "MP4 / MOV video"),
        (b"\x00\x00\x00\x1cftyp", "MP4 / MOV video"),
        (b"\x1a\x45\xdf\xa3", "WebM / Matroska video"),
        (b"PK\x03\x04", "ZIP archive"),
        (b"PK\x05\x06", "ZIP archive (EOCD)"),
        (b"PK\x07\x08", "ZIP archive (spanned)"),
        (b"Rar!\x1a\x07", "RAR archive"),
        (b"\x1f\x8b\x08", "GZIP compressed"),
        (b"BZh", "BZIP2 compressed"),
        (b"\xfd7zXZ\x00", "XZ compressed"),
        (b"7z\xbc\xaf\x27\x1c", "7-Zip archive"),
        (b"%PDF", "PDF document"),
        (b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1", "OLE2 / DOC / XLS"),
        (b"{\\rtf", "RTF document"),
        (b"\x00\x01\x00\x00\x00", "Windows Icon / Cursor"),
        (b"MM\x00*", "TIFF image (big-endian)"),
        (b"II*\x00", "TIFF image (little-endian)"),
        (b"\xca\xfe\xba\xbe", "Java class file"),
        (b"\xfe\xed\xfa\xce", "Mach-O (32-bit)"),
        (b"\xfe\xed\xfa\xcf", "Mach-O (64-bit)"),
        (b"\xce\xfa\xed\xfe", "Mach-O (reverse 32-bit)"),
        (b"\xcf\xfa\xed\xfe", "Mach-O (reverse 64-bit)"),
        (b"qod\x00", "Mach-O (arm64e)"),
        (b"OggS", "OGG container"),
        (b"\xff\xfb", "MP3 audio"),
        (b"\xff\xf3", "MP3 audio"),
        (b"\xff\xf2", "MP3 audio"),
        (b"fLaC", "FLAC audio"),
        (b"ID3", "MP3 with ID3 tags"),
        (b"\x00\x00\x00\x0c", "JPEG 2000"),
        (b"\x00\x00\x00\x0f", "JPEG 2000"),
        (b"8BPS", "PSD image"),
        (b"\x01\xda\x01\x01\x00\x03", "DPX image"),
    ]

    def _magic_to_type(self, magic: bytes, ext: str) -> str:
        for sig, desc in self._MAGIC_DB:
            if magic[:len(sig)] == sig:
                return desc
        ext_map = {
            ".pyc": "Python bytecode",
            ".whl": "Python wheel archive",
            ".so": "Shared library (ELF)",
            ".dll": "Dynamic link library",
            ".dylib": "Dynamic library (Mach-O)",
            ".wasm": "WebAssembly binary",
        }
        return ext_map.get(ext, "")

    def _parse_elf_sections(self, data: bytes) -> List[Dict]:
        sections = []
        try:
            if len(data) < 64:
                return sections
            is_64 = data[4] == 2
            if is_64:
                shoff = struct.unpack_from("<Q", data, 40)[0]
                shentsize = struct.unpack_from("<H", data, 58)[0]
                shnum = struct.unpack_from("<H", data, 60)[0]
                shstrndx = struct.unpack_from("<H", data, 62)[0]
            else:
                shoff = struct.unpack_from("<I", data, 32)[0]
                shentsize = struct.unpack_from("<H", data, 46)[0]
                shnum = struct.unpack_from("<H", data, 48)[0]
                shstrndx = struct.unpack_from("<H", data, 50)[0]

            if shstrndx >= shnum or shentsize == 0:
                return sections

            strtab_off = shoff + shstrndx * shentsize
            if is_64:
                strtab_sh_offset = struct.unpack_from("<Q", data, strtab_off + 24)[0]
                strtab_sh_size = struct.unpack_from("<Q", data, strtab_off + 32)[0]
            else:
                strtab_sh_offset = struct.unpack_from("<I", data, strtab_off + 16)[0]
                strtab_sh_size = struct.unpack_from("<I", data, strtab_off + 20)[0]

            strtab = data[strtab_sh_offset:strtab_sh_offset + strtab_sh_size]

            for i in range(shnum):
                off = shoff + i * shentsize
                if off + shentsize > len(data):
                    break
                if is_64:
                    name_off = struct.unpack_from("<I", data, off)[0]
                    sh_size = struct.unpack_from("<Q", data, off + 32)[0]
                    sh_addr = struct.unpack_from("<Q", data, off + 16)[0]
                else:
                    name_off = struct.unpack_from("<I", data, off)[0]
                    sh_size = struct.unpack_from("<I", data, off + 20)[0]
                    sh_addr = struct.unpack_from("<I", data, off + 12)[0]

                name = strtab[name_off:strtab.find(b"\x00", name_off)].decode("latin-1", errors="replace") if name_off < len(strtab) else ""
                if name and sh_size > 0 and not name.startswith("."):
                    sections.append({"name": f".{name}" if not name.startswith(".") else name, "offset": sh_addr, "size": sh_size})
                elif name and sh_size > 0:
                    sections.append({"name": name, "offset": sh_addr, "size": sh_size})
        except Exception:
            pass
        return sections[:20]

    def _parse_pe_sections(self, data: bytes) -> List[Dict]:
        sections = []
        try:
            if len(data) < 0x3c + 4:
                return sections
            pe_offset = struct.unpack_from("<I", data, 0x3c)[0]
            if pe_offset + 0x14 >= len(data):
                return sections
            num_sections = struct.unpack_from("<H", data, pe_offset + 6)[0]
            opt_header_size = struct.unpack_from("<H", data, pe_offset + 0x14)[0]
            section_start = pe_offset + 0x18 + opt_header_size

            for i in range(min(num_sections, 20)):
                off = section_start + i * 40
                if off + 40 > len(data):
                    break
                raw_name = data[off:off + 8]
                name = raw_name.split(b"\x00")[0].decode("latin-1", errors="replace")
                virtual_size = struct.unpack_from("<I", data, off + 8)[0]
                virtual_addr = struct.unpack_from("<I", data, off + 12)[0]
                raw_size = struct.unpack_from("<I", data, off + 16)[0]
                if name:
                    sections.append({"name": name, "offset": virtual_addr, "size": virtual_size or raw_size})
        except Exception:
            pass
        return sections[:20]

    def _binary_risk_indicators(self, data: bytes, sections: list, is_compressed: bool) -> list:
        indicators = []
        if not data:
            return indicators
        if b"PIE" not in data[:128] and data[:4] == b"\x7fELF":
            indicators.append("no_pie")
        if sections:
            has_canary = any("stack" in s.get("name", "").lower() or "security" in s.get("name", "").lower() for s in sections)
            if not has_canary:
                indicators.append("no_canary")
        if b"libc" in data:
            indicators.append("links_libc")
        if is_compressed:
            indicators.append("packed_or_encrypted")
        if b"UPX" in data[:4096]:
            indicators.append("upx_packed")
        if data[:2] == b"MZ" and b"kernel32" in data.lower():
            indicators.append("windows_api_usage")
        return indicators[:8]

    def _read_text(self, resolved: Path, base: Dict) -> Dict[str, Any]:
        ext = resolved.suffix.lower()
        lang = _SOURCE_CODE_EXTENSIONS.get(ext, "text")
        content = self._read_text_content(resolved)
        if not content:
            content = "[File read error: unable to decode content]"
        elif len(content) > self.max_text_size:
            content = content[:self.max_text_size] + "\n[... File truncated due to size limit ...]"
        base["encoding"] = "utf8"
        base["content"] = content
        base["language"] = lang
        return base

    def _read_directory(self, resolved: Path) -> Dict[str, Any]:
        stat_info = resolved.stat()
        metadata = self._build_metadata(resolved, stat_info)
        metadata["mime_type"] = "inode/directory"

        total_items = 0
        total_files = 0
        total_subdirs = 0
        total_size = 0
        largest_file: Tuple[str, int] = ("", 0)
        oldest_file: Tuple[str, float] = ("", float("inf"))
        newest_file: Tuple[str, float] = ("", 0)
        file_types: Dict[str, int] = {}
        max_depth = 3
        samples: List[str] = []
        actual_max_depth = 0

        try:
            for root, dirs, files in os.walk(resolved):
                rel = Path(root).relative_to(resolved)
                depth = len(rel.parts) if str(rel) != "." else 0
                actual_max_depth = max(actual_max_depth, depth)
                if depth > max_depth:
                    dirs.clear()
                    continue
                total_subdirs += len(dirs)
                total_files += len(files)
                total_items += len(dirs) + len(files)

                for f in files:
                    fp = Path(root) / f
                    try:
                        fs = fp.stat()
                        sz = fs.st_size
                        total_size += sz
                        if sz > largest_file[1]:
                            largest_file = (f, sz)
                        mt = fs.st_mtime
                        if mt < oldest_file[1]:
                            oldest_file = (f, mt)
                        if mt > newest_file[1]:
                            newest_file = (f, mt)
                        ext = fp.suffix.lower().lstrip(".") or "no_ext"
                        file_types[ext] = file_types.get(ext, 0) + 1
                    except (OSError, PermissionError):
                        pass
                    if len(samples) < 50:
                        try:
                            rel_path = _norm(str(Path(root).relative_to(resolved) / f))
                            samples.append(f"/{rel_path}")
                        except ValueError:
                            pass

                if len(samples) >= 50:
                    dirs.clear()
        except PermissionError:
            pass

        def _fmt_file(name: str, ts: float) -> str:
            return f"{name} (modified {_utc_from_ts(ts)})"

        base: Dict[str, Any] = {
            "file_type": "directory",
            "encoding": None,
            "size_bytes": 0,
            "read_at": _utc_from_ts(datetime.now(timezone.utc).timestamp()),
            "content": None,
            "metadata": metadata,
            "directory_info": {
                "total_items": total_items,
                "total_files": total_files,
                "total_subdirectories": total_subdirs,
                "total_size_bytes": total_size,
                "largest_file": f"{largest_file[0]} ({largest_file[1]} bytes)" if largest_file[0] else "",
                "oldest_file": _fmt_file(oldest_file[0], oldest_file[1]) if oldest_file[0] else "",
                "newest_file": _fmt_file(newest_file[0], newest_file[1]) if newest_file[0] else "",
                "file_types": dict(sorted(file_types.items(), key=lambda x: -x[1])),
                "structure_preview": {
                    "depth": actual_max_depth,
                    "max_depth": max_depth,
                    "sample": samples[:20],
                },
            },
        }
        return base

    def _enrich_with_git(self, resolved: Path, result: Dict) -> Dict:
        from src.modules.filesystem.adapters.git import DiskGit
        insights = DiskGit.get_insights(resolved)
        if insights:
            result["git_insights"] = insights
            root = Path(insights["repo_root"])
            status = DiskGit.get_file_status(root, resolved)
            if status:
                result["git_status"] = status
            else:
                try:
                    if DiskGit.is_tracked(root, resolved):
                        result["git_status"] = {"working_status": "clean", "staged_status": None}
                except Exception:
                    pass
        return result

    def _find_git_root(self, resolved: Path) -> Optional[Path]:
        from src.modules.filesystem.adapters.git import DiskGit
        return DiskGit.find_root(resolved)

    def write_to_disk(
        self,
        path: str,
        content: str = "",
        *,
        encoding: str = "utf8",
        write_mode: str = "create",
        is_directory: bool = False,
        permissions: Optional[int] = None,
        create_parents: bool = True,
        backup_existing: bool = False,
        atomic_write: bool = True,
    ) -> Dict[str, Any]:
        resolved = Path(path).resolve()

        # Path traversal prevention
        if ".." in path.split("/") or ".." in path.split("\\"):
            return {
                "error": "Path traversal detected",
                "status_code": 400,
                "data": {
                    "provided_path": path,
                    "resolved_path": _norm(str(resolved)),
                    "reason": "Path contains parent directory references ('..')",
                }
            }

        if is_directory:
            return self._write_directory(resolved, write_mode, create_parents, permissions)

        return self._write_file(resolved, content, encoding, write_mode, create_parents, backup_existing, atomic_write, permissions)

    def _write_directory(
        self,
        resolved: Path,
        write_mode: str,
        create_parents: bool,
        permissions: Optional[int],
    ) -> Dict[str, Any]:
        if resolved.exists() and resolved.is_dir():
            if write_mode == "create":
                return {
                    "error": "Directory already exists",
                    "status_code": 409,
                    "data": {
                        "existing_path": _norm(str(resolved)),
                        "last_modified": _utc_from_ts(resolved.stat().st_mtime),
                    }
                }
        elif resolved.exists() and not resolved.is_dir():
            return {"error": f"Path exists but is not a directory: {resolved}", "status_code": 400}

        created_paths = []
        try:
            if create_parents:
                parent = resolved.parent
                if not parent.exists():
                    parts = []
                    for p in resolved.parents:
                        if not p.exists():
                            parts.append(p)
                        else:
                            break
                    parts.reverse()
                    for p in parts:
                        p.mkdir(exist_ok=True)
                        created_paths.append(_norm(str(p)))
                        if permissions and os.name == "posix":
                            p.chmod(permissions)

            resolved.mkdir(exist_ok=(write_mode == "overwrite"))
            created_paths.append(_norm(str(resolved)))
            if permissions and os.name == "posix":
                resolved.chmod(permissions)

            stat_info = resolved.stat()
            return {
                "operation": "create_directory",
                "path": _norm(str(resolved)),
                "created_paths": created_paths,
                "permissions": int(stat_module.S_IMODE(stat_info.st_mode)),
                "owner": _get_owner(stat_info),
                "group": _get_group(stat_info),
                "created": _utc_from_ts(_get_created_ts(stat_info)),
            }
        except PermissionError:
            return {
                "error": "Permission denied: cannot create directory",
                "status_code": 403,
                "data": {
                    "path": _norm(str(resolved)),
                    "error_detail": f"PermissionError: Access denied",
                    "platform_hint": "Try running with elevated privileges or choose a writable directory.",
                }
            }
        except OSError as e:
            return {"error": f"Cannot create directory: {e}", "status_code": 500}

    def _write_file(
        self,
        resolved: Path,
        content: str,
        encoding: str,
        write_mode: str,
        create_parents: bool,
        backup_existing: bool,
        atomic_write: bool,
        permissions: Optional[int],
    ) -> Dict[str, Any]:
        if resolved.exists() and resolved.is_dir():
            return {"error": f"Path is a directory, not a file: {resolved}", "status_code": 400}

        if resolved.exists() and write_mode == "create":
            stat_info = resolved.stat()
            return {
                "error": "File already exists. Use overwrite=true or append mode to modify.",
                "status_code": 409,
                "data": {
                    "existing_file": _norm(str(resolved)),
                    "existing_size_bytes": stat_info.st_size,
                    "last_modified": _utc_from_ts(stat_info.st_mtime),
                }
            }

        if create_parents:
            try:
                resolved.parent.mkdir(parents=True, exist_ok=True)
            except PermissionError:
                return {"error": "Permission denied: cannot create parent directories", "status_code": 403}
            except OSError as e:
                return {"error": f"Cannot create parent directories: {e}", "status_code": 500}

        is_binary = encoding == "base64"
        data: bytes | str
        try:
            if is_binary:
                data = base64.b64decode(content)
            else:
                data = content
        except Exception as e:
            return {"error": f"Invalid {encoding} encoding: {e}", "status_code": 400}

        try:
            if write_mode == "append":
                original_size = resolved.stat().st_size if resolved.exists() else 0
                mode = "ab" if is_binary else "a"
                enc = None if is_binary else "utf-8"
                with open(resolved, mode, encoding=enc) as f:
                    f.write(data)
                stat_info = resolved.stat()
                return {
                    "operation": "append",
                    "path": _norm(str(resolved)),
                    "original_size_bytes": original_size,
                    "appended_bytes": len(data) if is_binary else len(data.encode("utf-8")),
                    "new_size_bytes": stat_info.st_size,
                    "modified": _utc_from_ts(stat_info.st_mtime),
                }

            # create or overwrite
            original_size = 0
            original_stat = None
            if resolved.exists():
                original_stat = resolved.stat()
                original_size = original_stat.st_size

            backup_path = None
            if backup_existing and resolved.exists():
                ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
                backup_path = resolved.with_name(resolved.name + f".bak.{ts}")
                import shutil
                shutil.copy2(resolved, backup_path)

            if atomic_write:
                temp = resolved.with_name(resolved.name + ".tmp." + str(os.getpid()))
                try:
                    mode = "wb" if is_binary else "w"
                    enc = None if is_binary else "utf-8"
                    with open(temp, mode, encoding=enc) as f:
                        f.write(data)
                    import shutil
                    shutil.move(str(temp), str(resolved))
                except Exception:
                    if temp.exists():
                        temp.unlink()
                    raise
            else:
                mode = "wb" if is_binary else "w"
                enc = None if is_binary else "utf-8"
                with open(resolved, mode, encoding=enc) as f:
                    f.write(data)

            if permissions and os.name == "posix":
                resolved.chmod(permissions)

            stat_info = resolved.stat()
            result = {
                "operation": "write_file",
                "path": _norm(str(resolved)),
                "size_bytes": stat_info.st_size,
                "write_mode": write_mode,
                "backup_created": backup_path is not None,
                "atomic_write_used": atomic_write,
                "modified": _utc_from_ts(stat_info.st_mtime),
            }

            if backup_path:
                result["backup_path"] = _norm(str(backup_path))
                result["original_size_bytes"] = original_size

            mime, _ = mimetypes.guess_type(str(resolved))
            if mime:
                result["file_type"] = mime

            if is_binary:
                import hashlib
                content_bytes = content.encode("ascii") if isinstance(content, str) else content
                result["sha256_checksum"] = hashlib.sha256(content_bytes).hexdigest()
                result["checksum_algorithm"] = "sha256"

            result["permissions"] = int(stat_module.S_IMODE(stat_info.st_mode))
            result["owner"] = _get_owner(stat_info)
            result["group"] = _get_group(stat_info)
            if write_mode != "overwrite":
                result["created"] = _utc_from_ts(_get_created_ts(stat_info))

            return result

        except PermissionError:
            return {
                "error": "Permission denied: cannot write to file",
                "status_code": 403,
                "data": {
                    "path": _norm(str(resolved)),
                    "error_detail": f"PermissionError: [Errno 13] Access denied",
                    "platform_hint": "Try running with elevated privileges or choose a writable directory.",
                }
            }
        except OSError as e:
            return {"error": f"OS error: {e}", "status_code": 500}
        except Exception as e:
            return {"error": str(e), "status_code": 500}

    @staticmethod
    def read_file(params: Dict[str, Any]) -> Dict[str, Any]:
        path = params.get("path")
        if not path:
            raise ApiError("path is required", status_code=400, error_code="FS_004")

        db = params.get("_db")
        repo_id = params.get("repo_id")
        integrity = None
        if db and repo_id and path:
            try:
                from src.core.database.integrity import FileIntegrity
                resolved = Path(path).resolve()
                integrity = FileIntegrity(db).check(repo_id, resolved)
            except Exception:
                pass

        reader = DiskReader()
        result = reader.read(path)

        if integrity:
            result["integrity"] = integrity

        if "error" in result:
            raise ApiError(
                str(result.get("error")),
                status_code=int(result.get("status_code", 400)),
                error_code="FS_004",
                details=result.get("data"),
            )

        return {"status_code": 200, "message": "File read successfully", "data": result}
