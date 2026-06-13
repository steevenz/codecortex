"""
Search.

:project: CodeCortex
:package: Modules.Filesystem.Adapters.Search
:author: Steeven Andrian
:copyright: (c) 2026 CODDY Codework
:standard: CODDY-Filesystem-v1.0
"""

from __future__ import annotations
from typing import Dict, Any, Optional, List, Tuple, Set
from pathlib import Path
import os
import re
import fnmatch
import difflib
from datetime import datetime, timezone
import uuid

from src.core import ApiError

BINARY_EXTENSIONS: Set[str] = {
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".ico", ".webp", ".svg",
    ".woff", ".woff2", ".ttf", ".eot", ".otf",
    ".mp3", ".mp4", ".avi", ".mov", ".wmv", ".flv", ".webm",
    ".zip", ".tar", ".gz", ".bz2", ".xz", ".7z", ".rar",
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".exe", ".dll", ".so", ".dylib", ".bin", ".deb", ".rpm",
    ".o", ".a", ".lib", ".obj",
    ".pyc", ".pyo", ".pyd",
    ".db", ".sqlite", ".sqlite3",
    ".ico", ".cur",
}

MAX_FILE_SIZE = 10 * 1024 * 1024
DEFAULT_MAX_RESULTS = 100
MAX_RESULTS_LIMIT = 1000
DEFAULT_CONTEXT_LINES = 2
MAX_MATCHES_PER_FILE = 50

EXTRACTABLE_EXTENSIONS: Set[str] = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx",
}


from src.core.utils.path import norm_path as _norm
from src.core.config.environment import utc_ts_to_iso as _utc_from_ts


def _is_binary(path: Path) -> bool:
    ext = path.suffix.lower()
    if ext in BINARY_EXTENSIONS:
        return True
    try:
        with open(path, "rb") as f:
            chunk = f.read(8192)
        return b"\x00" in chunk
    except Exception:
        return True


def _guess_language(path: Path) -> str:
    ext_map = {
        ".py": "python", ".js": "javascript", ".ts": "typescript",
        ".tsx": "typescriptreact", ".jsx": "javascriptreact",
        ".go": "go", ".rs": "rust", ".java": "java",
        ".cpp": "cpp", ".c": "c", ".h": "c", ".hpp": "cpp",
        ".cs": "csharp", ".php": "php", ".rb": "ruby",
        ".swift": "swift", ".kt": "kotlin", ".scala": "scala",
        ".lua": "lua", ".r": "r", ".m": "objectivec",
        ".pl": "perl", ".sql": "sql", ".sh": "shellscript",
        ".bash": "shellscript", ".zsh": "shellscript",
        ".md": "markdown", ".rst": "rst", ".txt": "text",
        ".json": "json", ".yaml": "yaml", ".yml": "yaml",
        ".toml": "toml", ".ini": "ini", ".cfg": "ini", ".conf": "ini",
        ".xml": "xml", ".html": "html", ".css": "css",
        ".scss": "scss", ".vue": "vue", ".svelte": "svelte",
        ".dockerfile": "dockerfile", ".gitignore": "ignore",
    }
    return ext_map.get(path.suffix.lower(), "")


class DiskSearch:
    def search(self, params: Dict[str, Any]) -> Dict[str, Any]:
        root_path = params.get("root_path", "")
        if not root_path:
            raise ApiError("root_path is required", status_code=400, error_code="FS_005")

        root = Path(root_path).resolve()
        if not root.exists():
            raise ApiError(f"Directory not found: {root_path}", status_code=404, error_code="FS_005")
        if not root.is_dir():
            raise ApiError("root_path must be a directory", status_code=400, error_code="FS_005")

        file_pattern = params.get("file_pattern", "*")
        file_regex_str = params.get("file_regex")
        content_regex_str = params.get("content_regex")
        flags_str = params.get("content_regex_flags", "")
        recursive = params.get("recursive", True)
        max_depth = params.get("max_depth")
        include_hidden = params.get("include_hidden", False)
        follow_symlinks = params.get("follow_symlinks", False)
        max_results = min(params.get("max_results", DEFAULT_MAX_RESULTS), MAX_RESULTS_LIMIT)
        include_snippet = params.get("include_content_snippet", True)
        context_lines = params.get("context_lines", DEFAULT_CONTEXT_LINES)
        replace_text = params.get("replace_text")
        dry_run = params.get("dry_run", True)

        exclude_patterns = params.get("exclude_patterns") or []

        if not file_pattern and not file_regex_str and not content_regex_str:
            raise ApiError("At least one of file_pattern, file_regex, or content_regex must be provided", status_code=400, error_code="FS_005")

        name_re: Optional[re.Pattern] = None
        if file_regex_str:
            try:
                name_re = re.compile(file_regex_str)
            except re.error as e:
                raise ApiError(f"Invalid file_regex: {e}", status_code=400, error_code="FS_005")

        content_re: Optional[re.Pattern] = None
        if content_regex_str:
            re_flags = 0
            if "i" in flags_str:
                re_flags |= re.IGNORECASE
            if "m" in flags_str:
                re_flags |= re.MULTILINE
            if "s" in flags_str:
                re_flags |= re.DOTALL
            try:
                content_re = re.compile(content_regex_str, re_flags)
            except re.error as e:
                raise ApiError(f"Invalid content_regex: {e}", status_code=400, error_code="FS_005")
        elif replace_text:
            raise ApiError("replace_text requires content_regex", status_code=400, error_code="FS_005")

        results: List[Dict[str, Any]] = []
        pending_replace: List[Dict[str, Any]] = []
        scanned = 0
        matched_files = 0
        skipped_binary = 0
        skipped_large = 0
        skipped_replace = 0
        hit_limit = False

        for current, dirs, files in os.walk(root, followlinks=follow_symlinks):
            depth = len(Path(current).relative_to(root).parts) if current != root else 0
            if max_depth is not None and depth > max_depth:
                del dirs[:]
                continue

            if not include_hidden:
                dirs[:] = [d for d in dirs if not d.startswith(".")]
                files = [f for f in files if not f.startswith(".")]

            dirs.sort(key=str.lower)
            files.sort(key=str.lower)

            for file in files:
                file_path = Path(current) / file
                scanned += 1

                if name_re:
                    if not name_re.search(file):
                        continue
                elif file_pattern != "*":
                    if not fnmatch.fnmatch(file, file_pattern):
                        continue

                if exclude_patterns:
                    rel = _norm(str(file_path.relative_to(root)))
                    if any(fnmatch.fnmatch(rel, pat) or fnmatch.fnmatch(file, pat) for pat in exclude_patterns):
                        continue

                stat_info = file_path.stat()
                file_size = stat_info.st_size

                if content_re:
                    if file_size > MAX_FILE_SIZE:
                        skipped_large += 1
                        continue

                    ext = file_path.suffix.lower()
                    if ext in EXTRACTABLE_EXTENSIONS:
                        text = self._extract_text(file_path)
                        if text is None:
                            skipped_binary += 1
                            continue
                        matches = self._search_in_text(text, content_re, include_snippet, context_lines)
                        if matches and replace_text:
                            skipped_replace += 1
                    elif _is_binary(file_path):
                        skipped_binary += 1
                        continue
                    else:
                        matches = self._search_content(file_path, content_re, include_snippet, context_lines)
                        if matches and replace_text:
                            pending_replace.append({
                                "path": str(file_path),
                                "content_re": content_re,
                                "replace_text": replace_text,
                            })

                    if not matches:
                        continue
                else:
                    matches = []

                matched_files += 1
                rel_path = _norm(str(file_path.relative_to(root)))
                entry: Dict[str, Any] = {
                    "file": {
                        "path": _norm(str(file_path)),
                        "relative_path": rel_path,
                        "size_bytes": file_size,
                        "modified": _utc_from_ts(stat_info.st_mtime),
                    },
                    "matches": matches,
                }

                lang = _guess_language(file_path)
                if lang:
                    entry["file"]["language"] = lang

                results.append(entry)

                if len(results) >= max_results:
                    hit_limit = True
                    break

            if hit_limit:
                break

        if not content_re and not content_regex_str:
            for entry in results:
                entry["matches"] = []

        replace_info: Optional[Dict[str, Any]] = None
        if replace_text:
            replace_info = self._do_replace(pending_replace, dry_run)

        data: Dict[str, Any] = {
            "query": {
                "root_path": _norm(str(root)),
                "file_pattern": file_pattern,
                "file_regex": file_regex_str,
                "content_regex": content_regex_str,
                "flags": flags_str,
                "recursive": recursive,
                "total_files_scanned": scanned,
                "total_matched_files": matched_files,
                "total_matches": sum(len(e["matches"]) for e in results),
            },
            "results": results,
        }

        if replace_info:
            data["replace_info"] = replace_info
        if skipped_binary:
            data["skipped_binary_count"] = skipped_binary
        if skipped_large:
            data["skipped_large_count"] = skipped_large
        if skipped_replace:
            data["skipped_replace_count"] = skipped_replace

        message = self._build_message(matched_files, skipped_binary, skipped_large, hit_limit, max_results, content_re is not None)
        status_code = 200 if not hit_limit else 206

        return {
            "status_code": status_code,
            "message": message,
            "data": data,
        }

    def _search_content(
        self,
        file_path: Path,
        content_re: re.Pattern,
        include_snippet: bool,
        context_lines: int,
    ) -> List[Dict[str, Any]]:
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                lines = f.readlines()
        except Exception:
            return []

        matches: List[Dict[str, Any]] = []
        for lineno, line in enumerate(lines, 1):
            for m in content_re.finditer(line):
                match_entry: Dict[str, Any] = {
                    "line": lineno,
                    "column": m.start(),
                    "snippet": line.rstrip("\n\r"),
                }

                if include_snippet:
                    before_start = max(0, lineno - context_lines - 1)
                    before_end = lineno - 1
                    if before_start < before_end:
                        match_entry["context_before"] = "".join(
                            lines[before_start:before_end]
                        ).rstrip("\n\r")
                    after_start = lineno
                    after_end = min(len(lines), lineno + context_lines)
                    if after_start < after_end:
                        match_entry["context_after"] = "".join(
                            lines[after_start:after_end]
                        ).rstrip("\n\r")

                matches.append(match_entry)
                if len(matches) >= MAX_MATCHES_PER_FILE:
                    return matches

        return matches

    def _extract_text(self, path: Path) -> Optional[str]:
        ext = path.suffix.lower()
        try:
            if ext == ".pdf":
                try:
                    from pypdf import PdfReader
                    reader = PdfReader(path)
                    pages = []
                    for page in reader.pages:
                        t = (page.extract_text() or "").strip()
                        if t:
                            pages.append(t)
                    return "\n".join(pages) if pages else None
                except ImportError:
                    return None

            elif ext in (".docx", ".doc"):
                try:
                    from docx import Document
                    doc = Document(path)
                    texts = [p.text for p in doc.paragraphs if p.text]
                    return "\n".join(texts) if texts else None
                except ImportError:
                    return None

            elif ext == ".xlsx":
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(path, read_only=True, data_only=True)
                    lines = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows(values_only=True):
                            lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                    return "\n".join(lines)
                except ImportError:
                    return None

            elif ext == ".xls":
                try:
                    from xlrd import open_workbook
                    wb = open_workbook(str(path))
                    lines = []
                    for sheet_name in wb.sheet_names():
                        ws = wb.sheet_by_name(sheet_name)
                        for row_idx in range(ws.nrows):
                            row_vals = [ws.cell_value(row_idx, col_idx) for col_idx in range(ws.ncols)]
                            lines.append(" | ".join(str(c) if c else "" for c in row_vals))
                    return "\n".join(lines)
                except ImportError:
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(path, read_only=True, data_only=True)
                        lines = []
                        for sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            for row in ws.iter_rows(values_only=True):
                                lines.append(" | ".join(str(c) if c is not None else "" for c in row))
                        return "\n".join(lines)
                    except ImportError:
                        return None
                    except Exception:
                        return None

        except Exception:
            return None
        return None

    def _search_in_text(
        self,
        text: str,
        content_re: re.Pattern,
        include_snippet: bool,
        context_lines: int,
    ) -> List[Dict[str, Any]]:
        lines = text.splitlines()
        matches: List[Dict[str, Any]] = []
        for lineno, line in enumerate(lines, 1):
            for m in content_re.finditer(line):
                match_entry: Dict[str, Any] = {
                    "line": lineno,
                    "column": m.start(),
                    "snippet": line,
                }
                if include_snippet:
                    before_start = max(0, lineno - context_lines - 1)
                    before_end = lineno - 1
                    if before_start < before_end:
                        match_entry["context_before"] = "\n".join(lines[before_start:before_end])
                    after_start = lineno
                    after_end = min(len(lines), lineno + context_lines)
                    if after_start < after_end:
                        match_entry["context_after"] = "\n".join(lines[after_start:after_end])
                matches.append(match_entry)
                if len(matches) >= MAX_MATCHES_PER_FILE:
                    return matches
        return matches

    def _do_replace(self, pending: List[Dict[str, Any]], dry_run: bool) -> Dict[str, Any]:
        affected: List[Dict[str, Any]] = []
        total_replaced = 0
        files_changed = 0
        files_failed = 0

        for item in pending:
            path = Path(item["path"])
            content_re = item["content_re"]
            replace_text = item["replace_text"]
            rel_path = _norm(str(path))

            try:
                old = path.read_text(encoding="utf-8", errors="replace")
            except Exception as e:
                affected.append({"path": rel_path, "status": "error", "error": str(e)})
                files_failed += 1
                continue

            new = content_re.sub(replace_text, old)
            if old == new:
                continue

            count = len(content_re.findall(old))
            total_replaced += count
            files_changed += 1

            diff = self._generate_diff(old, new, rel_path)

            affected.append({
                "path": rel_path,
                "match_count": count,
                "diff": diff,
                "status": "dry_run" if dry_run else "pending",
            })

            if not dry_run:
                try:
                    path.write_text(new, encoding="utf-8")
                    affected[-1]["status"] = "applied"
                except Exception as e:
                    affected[-1]["status"] = "error"
                    affected[-1]["error"] = str(e)
                    files_failed += 1

        result: Dict[str, Any] = {
            "mode": "dry_run" if dry_run else "apply",
            "total_replaced": total_replaced,
            "files_changed": files_changed,
            "files_failed": files_failed,
            "affected_files": affected,
        }

        return result

    def _generate_diff(self, old: str, new: str, rel_path: str) -> str:
        old_lines = old.splitlines(keepends=True)
        new_lines = new.splitlines(keepends=True)
        diff = difflib.unified_diff(
            old_lines, new_lines,
            fromfile=f"a/{rel_path}",
            tofile=f"b/{rel_path}",
            lineterm="",
        )
        return "".join(diff)

    def _build_message(
        self,
        matched_files: int,
        skipped_binary: int,
        skipped_large: int,
        hit_limit: bool,
        max_results: int,
        has_content_regex: bool,
    ) -> str:
        parts: List[str] = []
        if hit_limit:
            parts.append(f"Result limit reached ({max_results})")
            parts.append("Use more specific pattern or increase max_results")
        else:
            if matched_files == 0:
                parts.append("No files found matching the criteria")
            else:
                parts.append(f"Found {matched_files} file{'s' if matched_files != 1 else ''}")

        if skipped_binary:
            parts.append(f"{skipped_binary} binary file{'s' if skipped_binary != 1 else ''} skipped")
        if skipped_large:
            parts.append(f"{skipped_large} file{'s' if skipped_large != 1 else ''} over size limit skipped")

        return " | ".join(parts) if parts else "No results"

    def _respond(self, success: bool, status_code: int, message: str, data: Any = None) -> Dict[str, Any]:
        if not success or int(status_code) >= 400:
            raise ApiError(message, status_code=int(status_code), error_code="FS_005")
        return {"status_code": int(status_code), "message": message, "data": data}
